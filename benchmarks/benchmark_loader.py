import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import time
from datetime import datetime
import statistics
from openpyxl import Workbook, load_workbook

from cpu_metrics import CPUMetrics
from core.config import BenchmarksConfig

class BenchmarkCPUMetrics:
    """
    Classe per eseguire un benchmark CPU:
    - Esegue più snapshot (configurabili)
    - Calcola la media delle metriche numeriche
    - Gestisce liste e dizionari (medie)
    - Salva una sola riga finale in Excel
    """

    def __init__(self):
        self.sleep_time = BenchmarksConfig.SLEEP_TIME
        self.runs = BenchmarksConfig.RUNS
        self.excel_file = BenchmarksConfig.EXCEL_FILE
        self.cpu_metrics = CPUMetrics()

    def _average_dicts(self, dicts):
        """
        Calcola la media delle metriche numeriche in una lista di snapshot.
        - Liste per core → media totale singola
        - Dizionari (es: temperature) → media dei valori numerici
        - Altri tipi → lasciati invariati
        """

        result = {}
        keys = dicts[0].keys()

        for key in keys:
            first_value = dicts[0][key]

            # Valori numerici
            if isinstance(first_value, (int, float)):
                values = [d[key] for d in dicts if isinstance(d[key], (int, float))]
                result[key] = round(sum(values)/len(values), 2) if values else first_value

            # Liste (per core) → calcolo media totale
            elif isinstance(first_value, list):
                flat_values = []
                for d in dicts:
                    v = d[key]
                    if isinstance(v, list):
                        flat_values.extend([x for x in v if isinstance(x, (int, float))])
                result[key] = round(sum(flat_values)/len(flat_values), 2) if flat_values else None

            # Dizionari → media dei valori numerici
            elif isinstance(first_value, dict):
                nums = [x for x in first_value.values() if isinstance(x, (int, float))]
                result[key] = round(sum(nums)/len(nums), 2) if nums else None

            # Altri tipi → lascia come stringa
            else:
                result[key] = str(first_value)

        return result

    def _prepare_for_excel(self, data: dict) -> dict:
        """
        Garantisce che tutti i valori siano compatibili con Excel.
        Ripete la logica di _average_dicts come fallback
        """
        clean_data = {}
        for k, v in data.items():
            if isinstance(v, (int, float)):
                clean_data[k] = v
            elif isinstance(v, list):
                nums = [x for x in v if isinstance(x, (int, float))]
                clean_data[k] = round(sum(nums)/len(nums), 2) if nums else None
            elif isinstance(v, dict):
                nums = [x for x in v.values() if isinstance(x, (int, float))]
                clean_data[k] = round(sum(nums)/len(nums), 2) if nums else None
            else:
                clean_data[k] = str(v)
        return clean_data

    def _write_to_excel(self, data: dict):
        """
        Scrive i dati medi su Excel.
        - Se il file esiste → aggiunge una nuova riga
        - Se non esiste → crea file + intestazioni
        """
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(list(data.keys()))  # Intestazioni colonne

        ws.append(list(data.values()))
        wb.save(self.excel_file)
        print(f"Dati salvati in: {self.excel_file}")

    def run_benchmark(self):
        """
        Esegue il benchmark:
        - Esegue più snapshot (runs)
        - Pausa tra snapshot (sleep_time)
        - Calcola media finale
        - Salva su Excel
        """
        print(f"▶ Avvio benchmark CPU: {self.runs} rilevazioni, {self.sleep_time}s intervallo...")
        snapshots = []
        previous = None

        for i in range(self.runs):
            snap = self.cpu_metrics.snapshot(previous_snapshot=previous)
            snapshots.append(snap)
            previous = snap
            print(f"  Rilevazione {i+1}/{self.runs} completata")
            if i < self.runs - 1:
                time.sleep(self.sleep_time)

        # Media delle metriche
        avg_snapshot = self._average_dicts(snapshots)
        avg_snapshot["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Pulisce i valori per Excel
        avg_snapshot = self._prepare_for_excel(avg_snapshot)

        # Scrive su Excel
        self._write_to_excel(avg_snapshot)
        print("Benchmark completato.")

if __name__ == "__main__":
    benchmark = BenchmarkCPUMetrics()
    benchmark.run_benchmark()